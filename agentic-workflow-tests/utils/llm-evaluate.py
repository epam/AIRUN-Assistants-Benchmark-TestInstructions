#!/usr/bin/env python

import argparse
import csv
import logging
import math
import os
import re
import sys
import yaml

import numpy as NUMPY

from collections import defaultdict
from dataclasses import dataclass
from difflib import SequenceMatcher
from enum import Enum, StrEnum
from langchain_core.messages import HumanMessage
from langchain_openai import AzureChatOpenAI, ChatOpenAI
from openpyxl import Workbook
from openpyxl.styles import Font, NamedStyle, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.worksheet.table import Table, TableStyleInfo
from pathlib import Path
from typing import Self, TypeVar, Union
from xml.etree import ElementTree

from epam.auto_llm_eval import (
    evaluate_scenario,
    grade_scenario,
    read_file,
    write_file,
    CriteriaBase,
    CriterionEvalStep,
    CriterionEvalStepProcessed,
    CriteriaMeta,
    EvaluationStepsBucket,
    GradeResult,
)
from numpy import float64


logger = logging.getLogger(__name__)
logging.basicConfig(encoding='utf-8', level=logging.INFO)


def cdata(text: str) -> str:
    if re.search(r'[\<\&]', text):
        return f'<![CDATA[{text}]]>'
    return text


class WeightUnit(Enum):
    HIGH   = 1.0
    MEDIUM = 0.5
    LOW    = 0.2

    @classmethod
    def from_str(cls, s: str) -> Self:
        return cls[s.upper()]


class TestStage(StrEnum):
    FIRST = 'first'
    FINAL = 'final'


class TestStatus(Enum):
    PASSED = 1.0
    FAILED = 0.0

    def to_str(self):
        return self.name.capitalize()

    @classmethod
    def from_str(cls, status: str) -> Self:
        result = None
        match status:
            case 'Passed':
                result = TestStatus.PASSED
            case 'Failed':
                result = TestStatus.FAILED
            case _:
                raise ValueError(f'Invalid Assert type: {status}')
        return result

    @classmethod
    def from_bool(cls, status: bool) -> Self:
        result = TestStatus.PASSED if status else TestStatus.FAILED
        return result


class TestGradeField(StrEnum):
    NUMBER = 'Number'
    STAGE = 'Stage'
    TAG = 'Tag'
    NAME = 'Name'
    VALUE = 'Value'

    @classmethod
    def values(cls) -> tuple[str, ...]:
        return tuple(f.value for f in cls)


class TestGradeFieldName(StrEnum):
    ITERATION_COUNT = 'Iteration Count'
    PERFORMANCE = 'Performance'


@dataclass(frozen=True)
class TestGradeKey:
    number: str
    stage: TestStage
    tag: str


TestGradeValue = TypeVar("TestGradeValue", bound=Union[int, str, float, NUMPY.float64])


@dataclass
class TestGrades:
    __grades: dict[TestGradeKey, dict[str, TestGradeValue]]

    @staticmethod
    def to_fields_dict(key: TestGradeKey, field_name: str, field_value: str) -> dict[str, str]:
        fields = {
            TestGradeField.NUMBER: key.number,
            TestGradeField.STAGE: key.stage,
            TestGradeField.TAG: key.tag,
            TestGradeField.NAME: field_name,
            TestGradeField.VALUE: field_value,
        }
        return fields

    @classmethod
    def name_from_str(cls, name: str) -> TestGradeField | str:
        return TestGradeField(name) if name in (f.value for f in TestGradeField) else name

    @classmethod
    def value_from_str(cls, name: TestGradeField | str, value: str) -> TestGradeValue:
        return int(value) if name == TestGradeFieldName.ITERATION_COUNT else float(value)

    def write_to_csv(self, file_name: str):
        fieldnames = list(TestGradeField.values())
        with open(file_name, 'w', newline='\n') as csv_file:
            writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
            writer.writeheader()
            for key, value_dict in self.__grades.items():
                for name, value in value_dict.items():
                    fields = TestGrades.to_fields_dict(key, name, value)
                    writer.writerow(fields)


    @staticmethod
    def excel_graded_category_name(name: str, stage: TestStage) -> str:
        return f"{name}.{stage}";

    @staticmethod
    def excel_test_key(nbr: str, tag: str) -> tuple[str, str]:
        return (nbr, tag);

    def write_summary_to_excel(self, summary_excel_file: str) -> None:
        # Collect test grades table columns:
        # - fixed: Number, Tag
        # - TODO: add Language, Complexity
        # - fixed: TestGradeFieldName values
        # - named grade values for each TestStage
        # - test final grade
        #       Iteration Count    Performance    accuracy    completeness    Final accuracy    Final completeness    Grade
        test_table_headers = [
            TestGradeField.NUMBER.value, TestGradeField.TAG.value,
            TestGradeFieldName.ITERATION_COUNT.value, TestGradeFieldName.PERFORMANCE.value
        ]
        col_number_name = TestGradeField.NUMBER.value
        col_tag_name = TestGradeField.TAG.value
        col_tag_index = test_table_headers.index(col_tag_name) + 1
        col_tag_letter = get_column_letter(col_tag_index)
        col_iter_count_index = test_table_headers.index(TestGradeFieldName.ITERATION_COUNT.value) + 1
        col_iter_count_letter = get_column_letter(col_iter_count_index)
        col_perf_index = test_table_headers.index(TestGradeFieldName.PERFORMANCE.value) + 1
        col_perf_letter = get_column_letter(col_perf_index)

        test_nbrs = set()
        test_tags = set()
        test_categories = set()
        for key, value_dict in self.__grades.items():
            test_nbrs.add(key.number)
            test_tags.add(key.tag)
            for name in value_dict:
                test_categories.add(name)
        test_categories.remove(TestGradeFieldName.ITERATION_COUNT.value)
        test_categories.remove(TestGradeFieldName.PERFORMANCE.value)
        test_categories_list = list(test_categories)
        test_categories_list.sort()
        test_table_headers.extend([TestGrades.excel_graded_category_name(name, stage) for stage in TestStage for name in test_categories_list])
        col_grade_name = 'Grade'
        test_table_headers.append(col_grade_name)
        col_grade_index = len(test_table_headers)
        col_grade_letter = get_column_letter(col_grade_index)

        test_nbrs_list = list(test_nbrs)
        test_nbrs_list.sort()
        test_tags_list = list(test_tags)
        test_tags_list.sort()
        category_multiplier = 1.0 / len(test_categories_list) / 2

        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"

        total_style = NamedStyle(name="total")
        total_style.font = Font(bold=True, color="941BCC")
        total_style.fill = PatternFill(start_color="78E5F5", end_color="78E5F5", fill_type="solid")
        wb.add_named_style(total_style)

        # test grades table

        ws.append(test_table_headers)
        for nbr in test_nbrs_list:
            for tag in test_tags_list:
                test_first_key = TestGradeKey(nbr, TestStage.FIRST, tag)
                test_final_key = TestGradeKey(nbr, TestStage.FINAL, tag)
                row = [nbr, tag,
                       self.__grades[test_final_key][TestGradeFieldName.ITERATION_COUNT], self.__grades[test_final_key][TestGradeFieldName.PERFORMANCE]]
                grade_formulae = "="
                next_row = ws.max_row + 1
                next_column = len(row) + 1
                stage_plus = ""
                for key in [test_first_key, test_final_key]:
                    grade_formulae += f"{stage_plus}("
                    plus = ""
                    for category in test_categories_list:
                        row.append(self.__grades[key][category])
                        column_letter = get_column_letter(next_column)
                        grade_formulae += f"{plus}{column_letter}{next_row}"
                        plus = "+"
                        next_column += 1
                    perf_multiplier = f"*{col_perf_letter}{next_row}" if key == test_final_key else ""
                    grade_formulae += f")*{category_multiplier}{perf_multiplier}"
                    stage_plus = "+"

                row.append(grade_formulae)
                ws.append(row)

        table_tests_name = "Tests"
        table_tests = Table(displayName=table_tests_name, ref=f"A1:{col_grade_letter}{ws.max_row}")
        table_tests.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        ws.add_table(table_tests)

        total_average_row = ws.max_row + 1
        total_average_col = col_grade_index
        cell = ws.cell(row=total_average_row, column=total_average_col)
        cell.value=f"=AVERAGE({table_tests_name}[{col_grade_name}])"
        cell.style = total_style

        # per tag STDEV table

        table_run_stdev_name = "StdevPerTag"
        table_run_stdev_col_tag_name = 'Tag'
        table_run_stdev_col_tag_index = col_grade_index + 2
        table_run_stdev_col_tag_letter = get_column_letter(table_run_stdev_col_tag_index)
        table_run_stdev_col_grade_name = 'Grade'
        table_run_stdev_col_grade_index = table_run_stdev_col_tag_index + 1
        table_run_stdev_col_grade_letter = get_column_letter(table_run_stdev_col_grade_index)
        ws.cell(row=1, column=table_run_stdev_col_tag_index, value=table_run_stdev_col_tag_name)
        ws.cell(row=1, column=table_run_stdev_col_tag_index + 1, value=table_run_stdev_col_grade_name)
        table_run_stdev_last_row = len(test_tags_list) + 1
        table_run_stdev = Table(displayName=table_run_stdev_name, ref=f"{table_run_stdev_col_tag_letter}1:{table_run_stdev_col_grade_letter}{table_run_stdev_last_row}")
        table_run_stdev.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        ws.add_table(table_run_stdev)

        for row, tag in enumerate(test_tags_list, start=2):
            ws.cell(row=row, column=table_run_stdev_col_tag_index, value=tag)
            ws.cell(row=row, column=table_run_stdev_col_tag_index + 1,
                    value=f"=AVERAGEIF({table_tests_name}[{col_tag_name}],{table_run_stdev_col_tag_letter}{row},{table_tests_name}[{col_grade_name}])")
        cell = ws.cell(row=table_run_stdev_last_row + 1, column=table_run_stdev_col_tag_index)
        cell.value="Average"
        cell.style = total_style
        cell = ws.cell(row=table_run_stdev_last_row + 1, column=table_run_stdev_col_grade_index)
        cell.value=f"=AVERAGE({table_run_stdev_name}[{table_run_stdev_col_grade_name}])"
        cell.style = total_style
        cell = ws.cell(row=table_run_stdev_last_row + 2, column=table_run_stdev_col_tag_index)
        cell.value="STDEV.S"
        cell.style = total_style
        cell = ws.cell(row=table_run_stdev_last_row + 2, column=table_run_stdev_col_grade_index)
        cell.value=f"=_xlfn.STDEV.S({table_run_stdev_name}[{table_run_stdev_col_grade_name}])"
        cell.style = total_style

        # per test STDEV

        test_stdev_col_number_name = 'Number'
        test_stdev_col_number_index = table_run_stdev_col_grade_index + 2
        test_stdev_col_number_letter = get_column_letter(test_stdev_col_number_index)
        ws.cell(row=1, column=test_stdev_col_number_index, value=test_stdev_col_number_name)
        test_stdev_col_number_formula_text = f"=_xlfn.UNIQUE(_xlfn.TOCOL({table_tests_name}[{col_number_name}]))"
        test_stdev_latest_row_index = len(test_nbrs_list) + 1
        ws.cell(row=2, column=test_stdev_col_number_index,
                value=ArrayFormula(ref=f"{test_stdev_col_number_letter}2:{test_stdev_col_number_letter}{test_stdev_latest_row_index}",
                                   text=test_stdev_col_number_formula_text))

        test_stdev_row_tag_formula_text = f"=_xlfn.UNIQUE(_xlfn.TOROW({table_tests_name}[{col_tag_name}]))"
        test_stdev_tag_col_index = test_stdev_col_number_index + 1
        test_stdev_tag_col_letter = get_column_letter(test_stdev_tag_col_index)
        test_stdev_tag_latest_col_index = test_stdev_col_number_index + len(test_tags_list)
        test_stdev_tag_latest_col_letter = get_column_letter(test_stdev_tag_latest_col_index)
        ws.cell(row=1, column=test_stdev_tag_col_index,
                value=ArrayFormula(ref=f"{test_stdev_tag_col_letter}1:{test_stdev_tag_latest_col_letter}1",
                                   text=test_stdev_row_tag_formula_text))

        for row in range(2, test_stdev_latest_row_index + 1):
            for col in range(test_stdev_tag_col_index, test_stdev_tag_latest_col_index + 1):
                col_letter = get_column_letter(col)
                value_text = f"=_xlfn.XLOOKUP(1,({table_tests_name}[{col_tag_name}]={col_letter}1)*({table_tests_name}[{col_number_name}]={test_stdev_col_number_letter}{row}),{table_tests_name}[{col_grade_name}])"
                ws.cell(row=row, column=col, value=value_text)

        test_stdev_col_average_name = 'Average'
        test_stdev_col_average_index = test_stdev_tag_latest_col_index + 1
        test_stdev_col_average_letter = get_column_letter(test_stdev_col_average_index)
        cell = ws.cell(row=1, column=test_stdev_col_average_index)
        cell.value=test_stdev_col_average_name
        cell.style = total_style
        test_stdev_col_stdevs_name = 'STDEV.S'
        test_stdev_col_stdevs_index = test_stdev_tag_latest_col_index + 2
        test_stdev_col_stdevs_letter = get_column_letter(test_stdev_col_stdevs_index)
        cell = ws.cell(row=1, column=test_stdev_col_stdevs_index)
        cell.value=test_stdev_col_stdevs_name
        cell.style = total_style
        for row in range(2, test_stdev_latest_row_index+1):
            cell = ws.cell(row=row, column=test_stdev_col_average_index)
            cell.value=f"=AVERAGE({test_stdev_tag_col_letter}{row}:{test_stdev_tag_latest_col_letter}{row})"
            cell.style = total_style
            cell = ws.cell(row=row, column=test_stdev_col_stdevs_index)
            cell.style = total_style
            cell.value=f"=_xlfn.STDEV.S({test_stdev_tag_col_letter}{row}:{test_stdev_tag_latest_col_letter}{row})"
            cell.style = total_style

        # styling

        for cell in ws[col_iter_count_letter]:
            cell.number_format = '0'
        number_col_format = '0.00'
        for col in range(col_perf_index, len(test_table_headers) + 1):
            col_letter = get_column_letter(col)
            for cell in ws[col_letter]:
                cell.number_format = number_col_format
        for cell in ws[table_run_stdev_col_grade_letter]:
            cell.number_format = number_col_format

        for col in range(test_stdev_tag_col_index, test_stdev_col_stdevs_index + 1):
            col_letter = get_column_letter(col)
            for cell in ws[col_letter]:
                cell.number_format = number_col_format

        wb.save(summary_excel_file)

    @classmethod
    def read_from_csv(cls, file_name: str) -> Self:
        grades: dict[TestGradeKey, dict[str, TestGradeValue]] = defaultdict(dict)
        try:
            with open(file_name) as csv_file:
                reader = csv.DictReader(csv_file)
                for row in reader:
                    number = row[TestGradeField.NUMBER]
                    stage = row[TestGradeField.STAGE]
                    tag = row[TestGradeField.TAG]
                    name = cls.name_from_str(row[TestGradeField.NAME])
                    value = cls.value_from_str(name, row[TestGradeField.VALUE])
                    key = TestGradeKey(number, stage, tag)
                    grades[key][name] = value
        except FileNotFoundError:
            logger.warning(f"The file '{file_name}' was not found.")
        return cls(grades.copy())

    def upsert(self, from_test_grades: Self):
        for key, value in from_test_grades.__grades.items():
            self.__grades[key] = value.copy()

    @classmethod
    def summarize_test_grades(cls, test_grades_list: list[Self], summary_test_grades_file_name: str):
        #TODO: need to be refactored, relies on legacy data model
        first_test_grades_list = []
        final_test_grades_list = []
        grade_names = []
        for test_grades in test_grades_list:
            if test_grades.performance_grade:
                final_test_grades_list.append(test_grades)
            else:
                first_test_grades_list.append(test_grades)
            grade_names.extend(test_grades.__grades.keys())

        grade_names = list(set(grade_names))
        summary = {}
        summary_fieldnames = []
        for grade_name in grade_names:
            grades = []
            weights = []
            for test_grades in first_test_grades_list:
                grades.append(test_grades.__grades[grade_name])
                weights.append(WeightUnit.from_str(test_grades.complexity).value)
            data_values = NUMPY.array(grades)
            data_weights = NUMPY.array(weights)
            avg = NUMPY.average(data_values, weights=data_weights)
            fieldname = 'first_' + grade_name
            summary[fieldname] = avg
            summary_fieldnames.append(fieldname)

            grades = []
            weights = []
            for test_grades in final_test_grades_list:
                grades.append(test_grades.__grades[grade_name])
                weights.append(WeightUnit.from_str(test_grades.complexity).value)
            data_values = NUMPY.array(grades)
            data_weights = NUMPY.array(weights)
            avg = NUMPY.average(data_values, weights=data_weights)
            fieldname = 'final_' + grade_name
            summary[fieldname] = avg
            summary_fieldnames.append(fieldname)

        grades = []
        weights = []
        for test_grades in final_test_grades_list:
            grades.append(test_grades.performance_grade)
            weights.append(WeightUnit.from_str(test_grades.complexity).value)
        data_values = NUMPY.array(grades)
        data_weights = NUMPY.array(weights)
        avg = NUMPY.average(data_values, weights=data_weights)
        fieldname = 'performance'
        summary[fieldname] = avg
        summary_fieldnames.append(fieldname)

        #print(f'summary = {summary}')
        #print(f'summary_fieldnames = {summary_fieldnames}')
        with open(summary_test_grades_file_name, 'x', encoding='utf-8', newline='\n') as file:
            writer = csv.DictWriter(file, fieldnames=summary_fieldnames)
            writer.writeheader()
            writer.writerow(summary)


@dataclass
class Assert:
    condition: str
    weight: WeightUnit

    @classmethod
    def from_element(
        cls, element
    ) -> Self:
        condition = element.text
        weight = element.attrib.get('weight', WeightUnit.HIGH.name)

        return cls(
            condition = condition.strip(),
            weight = WeightUnit.from_str(weight)
        )


@dataclass
class Criterion:
    type: str
    weight: WeightUnit
    asserts: tuple[Assert]

    @classmethod
    def from_element(
        cls, element
    ) -> Self:
        type = element.attrib['type']
        weight = element.attrib.get('weight', WeightUnit.HIGH.name)
        asserts = element.findall('./Assert')

        return cls(
            type = type,
            weight = WeightUnit.from_str(weight),
            asserts = tuple(Assert.from_element(e) for e in asserts)
        )


@dataclass
class TestLogAssert:
    status: TestStatus
    reviwed_status: TestStatus
    confidence: float
    condition: str
    reasoning: str
    type: str

    @classmethod
    def from_element(
        cls, element
    ) -> Self:
        status = element.find('Status').text
        reviwed_status = element.find('ReviewedStatus').text
        confidence = float(element.find('Confidence').text)
        condition = element.find('Condition').text
        reasoning = element.find('Reasoning').text
        type = element.find('Type').text

        return cls(
            status = TestStatus.from_str(status),
            reviwed_status = TestStatus.from_str(reviwed_status) if reviwed_status else None,
            confidence = confidence,
            condition = condition.strip(),
            reasoning = reasoning.strip(),
            type = type
        )

    @classmethod
    def from_criterion_eval_step_processed(cls, step: CriterionEvalStepProcessed) -> Self:
        return cls(
            status = TestStatus.from_bool(step.passed),
            reviwed_status = None,
            confidence = step.confidence,
            condition = step.criterion,
            reasoning = step.explanation,
            type = step.type
        )

    def effective_status(self):
        return self.reviwed_status if self.reviwed_status else self.status


@dataclass
class TestLog:
    asserts: tuple[TestLogAssert, ...]

    @classmethod
    def from_grade_results(cls, grade_results: tuple[GradeResult, ...]) -> Self:
        asserts = []
        for result in grade_results:
            asserts.extend([TestLogAssert.from_criterion_eval_step_processed(step) for step in result.evaluation_steps])
        return cls(asserts=tuple(asserts))

    @classmethod
    def from_file(cls, file_name: str) -> Self:
        test_log_tree = ElementTree.parse(file_name)
        test_log = test_log_tree.getroot()
        test_log_assert_elements = test_log.findall('./Asserts/Assert')
        test_log_asserts = tuple(TestLogAssert.from_element(e) for e in test_log_assert_elements)
        return cls(asserts=test_log_asserts)

    def write_to_file(self, xml_file_name: str) -> None:
        with open(xml_file_name, 'x', encoding='utf-8', newline='\n') as f:
            xml1 = f'''<?xml version="1.0" encoding="UTF-8" standalone="yes" ?>
<TestLog>
    <Asserts>
'''
            f.write(xml1)
            for assert_ in self.asserts:
                condition = cdata(assert_.condition)
                status = assert_.status.to_str()
                reasoning = '' if assert_.reasoning is None else cdata(assert_.reasoning)
                f.write(f'''
        <Assert>
            <Condition>{condition}</Condition>
            <Type>{assert_.type}</Type>
            <Status>{status}</Status>
            <ReviewedStatus></ReviewedStatus>
            <ReviewCause></ReviewCause>
            <Confidence>{assert_.confidence}</Confidence>
            <Reasoning>
            {reasoning}
            </Reasoning>
        </Assert>
''')
            xml2 = f'''
    </Asserts>
</TestLog>
'''
            f.write(xml2)


@dataclass
class TestRefinement:
    comments: tuple[str, ...]

    @classmethod
    def from_file(cls, file_name: str) -> Self:
        test_ref_tree = ElementTree.parse(file_name)
        test_ref = test_ref_tree.getroot()
        test_ref_comment_elements = test_ref.findall('./Comment')
        ref_comments = tuple(c.text for c in test_ref_comment_elements)
        return cls(comments=ref_comments)


class TestSpec:
    __tree: ElementTree

    def __init__(self, test_spec_tree: ElementTree):
        self.__tree = test_spec_tree

    def _element_text(self, name: str) -> str:
        return self.__tree.getroot().find(name).text

    @property
    def id(self) -> str:
        return self._element_text('Id')

    @property
    def title(self) -> str:
        return self._element_text('Title')

    @property
    def category(self) -> str:
        return self._element_text('Category')

    @property
    def complexity(self) -> str:
        return self._element_text('Complexity')

    @property
    def code_repository(self) -> str:
        return self._element_text('CodeRepository')

    @property
    def criteria(self) -> tuple[Criterion, ...]:
        criterion_elements = self.__tree.getroot().findall('./Criteria/Criterion')
        return tuple(Criterion.from_element(e) for e in criterion_elements)

    @classmethod
    def from_file(cls, file_name: str) -> Self:
        test_spec_tree = ElementTree.parse(file_name)
        return cls(test_spec_tree)


class Criteria(CriteriaBase):

    def __init__(
        self, metadata: CriteriaMeta, criterion_eval_steps: tuple[CriterionEvalStep, ...]
    ):
        super().__init__(metadata, criterion_eval_steps)

    @classmethod
    def from_test_spec(cls, test_spec: TestSpec) -> Self:
        metadata = CriteriaMeta(
            category=test_spec.category,
            experiment="",
            repository=test_spec.code_repository,
            scenario_id=int(test_spec.id),
        )

        asserts = tuple((assert_, criterion.type) for criterion in test_spec.criteria for assert_ in criterion.asserts)
        criterion_eval_steps = tuple(CriterionEvalStep(assert_[0].condition, assert_[0].weight.value, assert_[1])
                                     for assert_ in asserts)

        return cls(metadata=metadata, criterion_eval_steps=criterion_eval_steps)

    def evaluation_steps_buckets(self):
        distinct_types = {step.type for step in self.criterion_eval_steps}
        for type_ in distinct_types:
            typed_evaluation_steps = [step for step in self.criterion_eval_steps if step.type == type_]
            yield EvaluationStepsBucket(name=type_, evaluation_steps=typed_evaluation_steps)


def match_assert_and_test_log_assert(cr_assert: Assert, test_log_asserts: list[TestLogAssert]) -> TestLogAssert | None:
    log_asserts = [(log_assert, SequenceMatcher(a=log_assert.condition, b=cr_assert.condition).ratio()) for log_assert in test_log_asserts]
    log_asserts = [tup for tup in log_asserts if tup[1] > 0.9]
    if not log_asserts:
        return None
    log_asserts = sorted(log_asserts, key=lambda tup: tup[1], reverse=True)
    return log_asserts[0][0]


def collect_criterion_test_log_asserts(criterion: Criterion, test_log_asserts: list[TestLogAssert]):
    log_asserts = []
    for cr_assert in criterion.asserts:
        log_assert = match_assert_and_test_log_assert(cr_assert, test_log_asserts)
        if not log_assert:
            raise ValueError(f'Not found test log assert with condition: {cr_assert.condition}')
        log_asserts.append(log_assert)
    return log_asserts


def calculate_criterion_grade(criterion: Criterion, test_log_asserts: list[TestLogAssert]):
    assert_statuses = []
    assert_weights = []
    for cr_assert in criterion.asserts:
        log_assert = match_assert_and_test_log_assert(cr_assert, test_log_asserts)
        if not log_assert:
            raise ValueError(f'Not found test log assert with condition: {cr_assert.condition}')
        status = log_assert.reviwed_status.value if log_assert.reviwed_status else log_assert.status.value
        weight = cr_assert.weight.value
        assert_statuses.append(status)
        assert_weights.append(weight)
    data_values = NUMPY.array(assert_statuses)
    data_weights = NUMPY.array(assert_weights)
    avg = NUMPY.average(data_values, weights=data_weights)
    return avg


def calculate_criteria_grades(test_spec: TestSpec, test_log: TestLog) -> dict[str, NUMPY.float64]:
    iCriterion = 0
    criterion_grades_weights = {}
    for criterion in test_spec.criteria:
        criterion_grade = calculate_criterion_grade(criterion, test_log.asserts)
        if logger.isEnabledFor(logging.INFO):
            criterion_log_asserts = collect_criterion_test_log_asserts(criterion, test_log.asserts)
            if logger.isEnabledFor(logging.INFO):
                asserts_text = '\n'.join(f'\t- {iAssert.effective_status().name}: {iAssert.condition}' for iAssert in criterion_log_asserts)
                logger.info(f"Criteria[{iCriterion}] grade = {criterion_grade}, type = '{criterion.type}':\n{asserts_text}")

        criterion_type = criterion.type
        criterion_grades_weights.setdefault(criterion_type, ([],[]))[0].append(criterion_grade)
        criterion_grades_weights[criterion_type][1].append(criterion.weight.value)

        iCriterion += 1

    criterion_grades = {}
    for criterion_type, grades_weights in criterion_grades_weights.items():
        data_values = NUMPY.array(grades_weights[0])
        data_weights = NUMPY.array(grades_weights[1])
        avg = NUMPY.average(data_values, weights=data_weights)
        criterion_grades[criterion_type] = avg
    return criterion_grades


def calculate_performance_grade(iteration_count: int) -> float | None:
    """
    Performance Grade Formula with Threshold for 1-2 Tries:
        Grade = M * math.exp(-a * max(0, T-s))
    where:
    M: Maximum grade (e.g., 1),
    T: Try count (number of attempts),
    a: Decay rate (controls how quickly the grade decreases after the 2nd try. e.g., 0.2),
    s: Shift parameter, which shifts the steep drop to occur after s attempts (set s=1 for this case).
    """
    T = iteration_count
    if T == 0:
        return 1.0
    M = 1.0
    a = 0.2
    s = 1
    grade = M * math.exp(-a * max(0, T-s))
    return grade


def get_eval_model():
    """Get the LLM model for generating evaluation reports."""
    # Read API key from the environment variables
    openai_api_key = os.environ["OPENAI_API_KEY"]
    if not openai_api_key:
        raise ValueError("OPENAI_API_KEY environment variable is not set")

    model = ChatOpenAI(
        model_name="o3-mini",
        temperature=1,
        api_key=openai_api_key,
    )

# TODO: change model. See https://github.com/epam/AIRUN-LLM-Benchmark/blob/eade4227edf889c8ce303259b189d6e6b19efbb3/Utils/llm/config.py#L60
#        "api_key": openai_api_key,
#        "max_tokens": max_tokens,
#        "model_id": model,
#        "temperature": 1,
#        "reasoning_effort": effort,
#        "verbosity": verbosity,
#        "background": background,

    return model


def evaluate(test_spec: str, evaluated_file: str, llm_report_dir: str, test_log_file: str) -> None:
    logger.info(f"Evalute file '{evaluated_file}' with test spec '{test_spec}', dump LLM reports to '{llm_report_dir}', write test log to '{test_log_file}'")

    if Path(test_log_file).exists():
        raise ValueError(f"File '{test_log_file}' exists. Could not rewrite.")

    test_spec = TestSpec.from_file(test_spec)

    criteria = Criteria.from_test_spec(test_spec)

    evaluated_content = read_file(Path(evaluated_file))

    eval_model = get_eval_model()

    def extract_json_from_md(content: str) -> str:
        json_text = content
        json_text = json_text.strip("\n")
        json_text = json_text.strip("`")
        json_text = json_text.replace("json\n", "", 1)

        return json_text

    def execute_prompt(prompt: str) -> str:
        message: HumanMessage = HumanMessage(content=prompt)
        api_response = eval_model.invoke([message])
        report: str = str(api_response.content)

        return extract_json_from_md(report)

    evaluation_results = evaluate_scenario(
        criteria=criteria,
        output=evaluated_content,
        execute_prompt=execute_prompt,
    )

    for result in evaluation_results:
        file_name = f"{result.name}.json"
        write_file(
            os.path.join(llm_report_dir, file_name),
            result.report,
        )

    grade_results = grade_scenario(evaluation_results)
    test_log = TestLog.from_grade_results(grade_results)
    test_log.write_to_file(test_log_file)


def grade(test_spec: str, test_log_file: str, test_stage: TestStage, test_refinement_file: str, test_report_file: str, test_tag: str) -> None:
    logger.info(f"Grade test evaluation results from '{test_log_file}' with test spec '{test_spec}', test stage = '{test_stage}', write test report to '{test_report_file}', tag with '{test_tag}'")

    test_spec = TestSpec.from_file(test_spec)
    test_log = TestLog.from_file(test_log_file)

    iteration_count = len(TestRefinement.from_file(test_refinement_file).comments) if test_refinement_file else None
    performance_grade = calculate_performance_grade(iteration_count) if iteration_count is not None and iteration_count>=0 else None

    criteria_grades = calculate_criteria_grades(test_spec, test_log)

    existing_test_grades = TestGrades.read_from_csv(test_report_file)

    key = TestGradeKey(number=test_spec.id, stage=test_stage, tag=test_tag)
    grades: dict[TestGradeKey, dict[str, TestGradeValue]] = defaultdict(dict)
    grades[key] = criteria_grades.copy()
    if iteration_count is not None and iteration_count>=0:
        grades[key][TestGradeFieldName.ITERATION_COUNT] = iteration_count
    if performance_grade:
        grades[key][TestGradeFieldName.PERFORMANCE] = performance_grade

    logger.info(f'Criteria grades = {criteria_grades}')
    logger.info(f'Iteration_count = {iteration_count}')
    logger.info(f'Performance grade = {performance_grade}')

    new_test_grades = TestGrades(grades)
    existing_test_grades.upsert(new_test_grades)
    existing_test_grades.write_to_csv(test_report_file)


def summarize_to_excel(test_report_file: str, summary_excel_file: str):
    logger.info(f"Summarize results from test report '{test_report_file}' to '{summary_excel_file}'")

    test_grades = TestGrades.read_from_csv(test_report_file)
    test_grades.write_summary_to_excel(summary_excel_file)


def main():
    ### common arguments parsers ###

    test_spec_parser = argparse.ArgumentParser(add_help=False)
    test_spec_parser.add_argument(
        "--test-spec",
        type=str,
        required=True,
        help="Test specification file",
    )

    evaluated_file_parser = argparse.ArgumentParser(add_help=False)
    evaluated_file_parser.add_argument(
        "--evaluated-file",
        type=str,
        required=True,
        help="Evaluated file (in Markdown format)",
    )

    llm_report_dir_parser = argparse.ArgumentParser(add_help=False)
    llm_report_dir_parser.add_argument(
        "--llm-report-dir",
        type=str,
        required=True,
        help="LLM reports directory",
    )

    test_log_file_parser = argparse.ArgumentParser(add_help=False)
    test_log_file_parser.add_argument(
        "--test-log",
        type=str,
        required=True,
        help="Test log file (in XML format)",
    )

    test_report_parser = argparse.ArgumentParser(add_help=False)
    test_report_parser.add_argument(
        "--test-report",
        type=str,
        required=True,
        help="Test report file (in CSV format)",
    )

    test_grade_parser = argparse.ArgumentParser(add_help=False)
    test_grade_parser.add_argument(
        "--test-stage",
        type=TestStage,
        choices=list(TestStage),
        required=True,
        help="Test stage to grade: %(choices)s",
    )
    test_grade_parser.add_argument(
        "--test-tag",
        type=str,
        required=False,
        help="Tag to mark the test in the report",
    )

    test_refinement_parser = argparse.ArgumentParser(add_help=False)
    test_refinement_parser.add_argument(
        "--test-refinement",
        type=str,
        required=False,
        help="Test refinement file (in XML format)",
    )

    summary_excel_parser = argparse.ArgumentParser(add_help=False)
    summary_excel_parser.add_argument(
        "--summary-excel",
        type=str,
        required=True,
        help="Test summary file (in Excel .xlsx format)",
    )


    ### main parsers ###

    main_parser = argparse.ArgumentParser(
        description="Evaluate benchmark scenarios with LLM."
    )

    main_subparsers = main_parser.add_subparsers(dest='command', required=True)

    evaluate_parser = main_subparsers.add_parser(
        'evaluate',
        parents=[test_spec_parser, evaluated_file_parser, llm_report_dir_parser, test_log_file_parser])

    evaluate_parser = main_subparsers.add_parser(
        'grade',
        parents=[test_spec_parser, test_log_file_parser, test_report_parser, test_grade_parser, test_refinement_parser])

    summarize_parser = main_subparsers.add_parser(
        'summarize',
        parents=[test_report_parser])

    summarize_subparsers = summarize_parser.add_subparsers(dest='summarize_command', required=True)

    summarize_subparsers.add_parser(
        'excel',
        parents=[summary_excel_parser] )

    args = main_parser.parse_args()

    ### processing ###

    command = args.command

    match command:
        case 'evaluate':
            evaluate(args.test_spec, args.evaluated_file, args.llm_report_dir, args.test_log)
        case 'grade':
            grade(args.test_spec, args.test_log, args.test_stage, args.test_refinement, args.test_report, args.test_tag)
        case 'summarize':
            summarize_command = args.summarize_command
            match summarize_command:
                case 'excel':
                    summarize_to_excel(args.test_report, args.summary_excel)
                case _:
                    logger.error(f"Invalid the command '{command}' sub-command '{summarize_command}'")
                    sys.exit(1)
        case _:
            logger.error(f'Invalid command: {command}')
            sys.exit(1)


if __name__ == "__main__":
    main()